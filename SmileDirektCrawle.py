import webbrowser
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import progressbar
import time
from selenium.webdriver.common.keys import Keys
import openpyxl
from selenium.webdriver.support.ui import Select
import signal
import sys

def signal_handler(signal, frame):
        print('You pressed Ctrl+C!')
        sys.exit(0)
signal.signal(signal.SIGINT, signal_handler)
print('Press Ctrl+C')

#-------------- Starts: Load Excel ------------------------
wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename='smile main sample.xlsm')
sheets = wb.sheetnames
ws = wb[sheets[0]]

#-------------- Ends: Load Excel ------------------------
successBar = progressbar.ProgressBar(maxval=31, \
    widgets=[progressbar.Bar('=', 'Success bar: [', ']'), ' ', progressbar.Percentage()])


successBar.start()

for i in range(2,31):
	index = str(i)
	successBar.update(i)
	# ------------- Starts: Reading Excel -------------------
	sdId = ws['A'+index].value
	brand = ws['P'+index].value
	model = ws['T'+index].value+" "+ws['U'+index].value
	inverhrkersetzung = ws['V'+index].value
	accessories = ws['R'+index].value
	leasing = ws['H'+index].value
	gender = ws['J'+index].value
	use = ws['I'+index].value
	licenceAge = ws['K'+index].value
	nat = ws['L'+index].value
	bd = '01.01.1976'
	deductibleTeil = ws['N'+index].value
	deductibleVoll = ws['O'+index].value
	
	licenceExperience = 2017-licenceAge
	
	if licenceExperience == 4:
		bd = '01.01.1994'
	elif licenceExperience < 5:
		bd = '01.01.1996'
		
	
	zipcode = ws['M2'].value

	
	#-------------- Ends: Reading Excel ---------------------
	
	try:
		driver = webdriver.Chrome()
		driver.set_window_size(1280, 1060)
	
		driver.get("https://app.smile-direct.ch/car?lang=de&sac=WOA")

		
		driver.find_element_by_xpath("//label[@for='markeundtyp_radio']").click()

		time.sleep(1)

		driver.find_element_by_xpath("//*[@id='brand']/option[text()='"+brand+"']").click()
		time.sleep(1)
	
		driver.find_element_by_xpath("//*[@id='type']/option[text()='"+model+"']").click()
		time.sleep(1)
	
	

		driver.find_element_by_xpath("//*[@id='inverkehrssetzungsJahr']/option[text()='2017']").click()
		driver.find_element_by_xpath("//*[@id='neuwertZubehoer']").send_keys(int(accessories))
		time.sleep(1)
		rowNumbers = len(driver.find_elements_by_xpath("//*[@id='modellist_tbody']/tr"))
		driver.find_element_by_xpath("//*[@id='modellist_tbody']/tr["+str(rowNumbers)+"]/td[1]/label").click()
		time.sleep(1)
		if leasing == 'yes':
			driver.find_element_by_xpath("//label[@for='finanzierung-leasing']").click()
		time.sleep(1)
		if gender == 'Male':
			driver.find_element_by_xpath("//*[@id='smiForm']/div[3]/div/div/div/ul[2]/li[10]/div[2]/div/label[1]").click()
		elif gender == 'Female':
			driver.find_element_by_xpath("//*[@id='smiForm']/div[3]/div/div/div/ul[2]/li[10]/div[2]/div/label[2]").click()
		time.sleep(1)
		driver.find_element_by_xpath("//*[@id='fahrzeugLenkerGeburtsdatum']").send_keys(bd)
		time.sleep(1)	
		driver.find_element_by_xpath("//*[@id='nationLenker']/option[text()='"+nat+"']").click()

		if nat != 'Schweiz':
			driver.find_element_by_xpath("//*[@id='bewilligungLenkerLabel']/option[text()='C']").click()
		driver.find_element_by_xpath("//*[@id='fahrzeugLenker.postleitzahl']").send_keys(int(zipcode))
		time.sleep(1)
		driver.find_element_by_xpath("//*[@id='anzahlKinder']").send_keys(0)
		time.sleep(1)
		driver.find_element_by_xpath("//*[@id='datumFahrpruefungPkw']").send_keys(int(licenceAge))

		if use == 'Private':
			driver.find_element_by_xpath("//*[@id='car_private_label']").click()
		elif use == 'Commute':	
			driver.find_element_by_xpath("//*[@id='car_privateway_label']").click()
		elif use == 'Business':
			driver.find_element_by_xpath("//*[@id='car_privatebusiness_label']").click()
		time.sleep(1)
		driver.find_element_by_xpath("//*[@id='smiForm']/div[3]/div/div/div/div/li/button").click()
		#wait.until(EC.element_located_to_be_selected((By.ID,'smiForm')))
		#----------------------------------------------------------------------------------------------------
		driver.find_element_by_xpath("//*[@id='smiForm']/div[2]/div/ul/li[1]/div[1]/div/label[2]").click()
		driver.find_element_by_xpath("//*[@id='smiForm']/div[2]/div/ul/li[3]/div[1]/div/label[2]").click()
		driver.find_element_by_xpath("//*[@id='smiForm']/div[2]/div/ul/li[5]/div[1]/div/label[2]").click()
		driver.find_element_by_xpath("//*[@id='schaeden']/div[1]/div/label[2]").click()
		driver.find_element_by_xpath("//*[@id='weiter']").click()
		#-----------------------------------------------------------------------------------------------------
		if leasing == 'no':
			time.sleep(2)
			driver.find_element_by_xpath("//*[@id='haftpflicht_content']/table/tbody/tr[2]/td[2]/div/div/div/div/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div/div/div/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='insassenunfall_acc']/li/div/div/div/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='gfverzicht_mod']/div/ul/li/div/div/div/div/label[1]").click()
			time.sleep(5)
			firstValue = driver.find_element_by_xpath("//*[@id='bruttopraemie']").text
			
			
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[1]/div/div/label[2]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_inner']/a").click()
			time.sleep(5)
			
			if deductibleTeil == 300:
				driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[5]/td[2]/div/div/div/div[1]/label[3]").click()
			elif deductibleTeil == 0:
				driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[5]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[17]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			secondValue = driver.find_element_by_xpath("//*[@id='bruttopraemie']").text
			
		
		
	
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[1]/div/div/label[3]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[2]/td[2]/div/div/div/div/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[4]/td[2]/div/div/div/div[1]/label[1]").click()
			
			if deductibleTeil == 300:

				driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[5]/td[2]/div/div/div/div[1]/label[3]").click()
			elif deductibleTeil == 0:
				driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[5]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[16]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[17]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			thirdValue = driver.find_element_by_xpath("//*[@id='bruttopraemie']").text
			
			ws['W'+index] = firstValue
			ws['X'+index] = secondValue
			ws['Y'+index] = thirdValue
		else:
			
			driver.find_element_by_xpath("//*[@id='haftpflicht_content']/table/tbody/tr[2]/td[2]/div/div/div/div/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_inner']/a").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[1]/div/div/label[3]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[2]/td[2]/div/div/div/div/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[4]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			if deductibleTeil == 300:
				driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[5]/td[2]/div/div/div/div[1]/label[3]").click()
				
			elif deductibleTeil == 0:
				driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[5]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[16]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='insassenunfall_acc']/li/div/div/div/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='kasko_acc']/li/div[2]/table/tbody/tr[17]/td[2]/div/div/div/div[1]/label[1]").click()
			time.sleep(5)
			driver.find_element_by_xpath("//*[@id='assistance_acc']/li/div[1]/div/div/label[1]").click()
			driver.find_element_by_xpath("//*[@id='gfverzicht_mod']/div/ul/li/div/div/div/div/label[1]").click()
			thirdValue = driver.find_element_by_xpath("//*[@id='bruttopraemie']").text
			
			ws['Y'+index] = thirdValue
		
		
	except:
		#print "An error occured at: "+str(sdId)
		
		output = open('log.txt','a')
		output.write("Failed record with ID: "+str(sdId)+"\n")
		output.close()
		wb.save("smile main sample.xlsm")
		pass
	finally:
		
		driver.close()
		driver.quit()
wb.save("smile main sample.xlsm")
successBar.finish()












